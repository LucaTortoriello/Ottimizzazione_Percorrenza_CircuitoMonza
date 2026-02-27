"""
GENERAZIONE DATASET MONZA
Questo script esegue uno sweep (scansione) di due parametri:
1. mu: Coefficiente di aderenza (grip) degli pneumatici.
2. W: Peso del consumo di carburante nella funzione obiettivo (strategia di gara).
Per ogni combinazione, ottimizza la velocità nei vari tratti del circuito per minimizzare
il tempo sul giro e il consumo, salvando i risultati in un file Excel.
"""
import numpy as np
from scipy.optimize import minimize
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings

# Disabilita gli avvisi di warning 
warnings.filterwarnings('ignore')

# --- 1. DATI DEL CIRCUITO DI MONZA ---
# Lunghezze totali dei 10 tratti del circuito (metri)
L = np.array([324.33, 458.00, 687.00, 412.00, 229.00,
              229.00, 916.00, 916.00, 705.67, 916.00])

# Spazio dedicato alla frenata per ogni tratto (metri)
x_brake = np.array([122.0, 0.0, 0.0, 107.0, 0.0, 35.0, 50.0, 104.0, 76.0, 0.0])

# Distanza standard di accelerazione/frenata per i tratti veloci
dist_frenata = 380.0

# Spazio dedicato all'accelerazione per ogni tratto (metri)
x_acc = np.array([202.33, dist_frenata, dist_frenata, 305.0, 229.0,
                  194.0, dist_frenata, dist_frenata, dist_frenata, dist_frenata])

# Spazio percorso a velocità costante
x_const = L - x_acc - x_brake

# --- 2. PARAMETRI FISICI DELLA VETTURA ---
m = 722.0       # Massa della vettura (kg)
g = 9.81        # Accelerazione di gravità (m/s^2)
k_aero = 0.387  # Coefficiente di resistenza aerodinamica
epsilon = 6e-6  # Coefficiente di consumo specifico (litri per velocità^2)

def build_vlim(mu):
    """
    Calcola le velocità limite di percorrenza in curva in base al grip (mu).
    Applica una formula fisica semplificata per le curve che dipendono dal grip.
    """
    rho = 224.62
    beta = np.sqrt(mu * rho * g) * 3.6  # Velocità limite calcolata in km/h
    # Ritorna array in m/s
    return np.array([90, 315, 135, 195, 170, 325, 185, beta, beta, 360.0]) / 3.6

def obj(v_peak, xa, xc, xb, v_lim, W):
    """
    FUNZIONE OBIETTIVO: J = Tempo + W * Consumo
    Minimizzare questa funzione significa trovare il miglior compromesso tra velocità e risparmio.
    """
    T = 0.0; C = 0.0; v_in = 0.1  # Inizializzazione tempo, consumo e velocità iniziale
    for i in range(len(v_peak)):
        vp = v_peak[i]; vt = v_lim[i] # v_peak è la variabile da ottimizzare, vt è il limite del tratto
        
        # Calcolo del Tempo nel tratto (Accelerazione + Costante + Frenata)
        T += xa[i]/((v_in+vp)/2) + xc[i]/vp + xb[i]/((vp+vt)/2)
        
        # Calcolo del Consumo nel tratto (basato sul quadrato della velocità media nelle fasi)
        v_a = ((v_in+vp)/2)*3.6; v_c = vp*3.6; v_b = ((vp+vt)/2)*3.6
        C += epsilon*(v_a**2)*(xa[i]/1000) + epsilon*(v_c**2)*(xc[i]/1000) + epsilon*(v_b**2)*(xb[i]/1000)
        
        v_in = vt # La velocità finale del tratto diventa l'entrata del successivo
    return T + W * C

def build_constraints(xa, xb, v_lim, mu):
    """
    Costruisce i vincoli di fattibilità fisica: la vettura non può accelerare 
    più di quanto il grip e l'aerodinamica permettano.
    """
    F_grip = mu * m * g
    cons = []
    def make_con(i):
        def c(v):
            v_in = 0.1 if i == 0 else v_lim[i-1]
            vp = v[i]; vt = v_lim[i]
            # Forza disponibile (Grip - Resistenza Aero)
            F_acc = F_grip - k_aero*((v_in+vp)/2)**2
            F_brk = F_grip + k_aero*((vp+vt)/2)**2
            # Vincoli energetici (Lavoro della forza >= Variazione energia cinetica)
            c1 = F_acc*xa[i] - 0.5*m*(vp**2 - v_in**2)
            c2 = F_brk*xb[i] - 0.5*m*(vp**2 - vt**2)
            return np.array([c1, c2])
        return c
    for i in range(len(xa)):
        cons.append({'type': 'ineq', 'fun': make_con(i)})
    return cons

def solve(mu, W):
    """
    Risolve il problema di ottimizzazione per una specifica coppia mu e W.
    """
    v_lim = build_vlim(mu)
    lb = v_lim.copy()                     # Limite inferiore: velocità minima in curva
    ub = np.full(len(lb), 360/3.6)        # Limite superiore: 360 km/h
    v0 = np.clip(v_lim + 10, lb, ub)      # Guess iniziale
    bounds = list(zip(lb, ub))            # Confini per le variabili
    cons = build_constraints(x_acc, x_brake, v_lim, mu) # Vincoli fisici
    
    # Algoritmo SLSQP per la minimizzazione vincolata
    res = minimize(obj, v0, args=(x_acc, x_const, x_brake, v_lim, W),
                   method='SLSQP', bounds=bounds, constraints=cons,
                   options={'ftol':1e-9, 'maxiter':2000, 'disp':False})
    
    # Calcolo del tempo giro puro (senza il peso W) dal risultato dell'ottimizzazione
    vp = res.x; v_in = 0.1; T = 0.0
    for i in range(len(vp)):
        vt = v_lim[i]
        T += x_acc[i]/((v_in+vp[i])/2) + x_const[i]/vp[i] + x_brake[i]/((vp[i]+vt)/2)
        v_in = vt
    return T

# --- 3. ESECUZIONE DELLO SWEEP (GENERAZIONE DATI) ---
mu_vals = np.round(np.arange(1.0, 3.6, 0.2), 2)
W_vals = np.round(np.arange(0.4, 20.2, 0.2), 2)

total = len(mu_vals) * len(W_vals)
print(f"Calcolo {total} combinazioni (mu x W) ...")

results = []
for idx_mu, mu in enumerate(mu_vals):
    for idx_W, W in enumerate(W_vals):
        T_s = solve(mu, W)     # Esegue l'ottimizzazione
        T_m = T_s / 60.0       # Converte in minuti
        results.append((mu, W, T_s, T_m)) # Salva in memoria
        
        # Feedback visivo nel terminale ogni 50 iterazioni
        done = idx_mu*len(W_vals) + idx_W + 1
        if done % 50 == 0 or done == total:
            print(f"  {done}/{total}  mu={mu:.1f}  W={W:.1f}  T={T_s:.3f}s")

# --- 4. FORMATTAZIONE E SALVATAGGIO EXCEL ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Dataset"

# Definizione stili grafici per Excel
hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill('solid', start_color='1F4E79') # Blu scuro
alt_fill = PatternFill('solid', start_color='D6E4F0') # Blu chiaro (righe alternate)
num_fmt_s = '0.000'
num_fmt_m = '0.00000'
center = Alignment(horizontal='center', vertical='center')
thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Scrittura intestazioni
headers = ['mu [-]', 'W [s/L]', 'Tempo [s]', 'Tempo [min]']
col_widths = [12, 12, 14, 14]

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = center; cell.border = border
    ws.column_dimensions[get_column_letter(col)].width = w

# Scrittura dei dati calcolati
for row_idx, (mu, W, Ts, Tm) in enumerate(results, start=2):
    fill = alt_fill if row_idx % 2 == 0 else None # Colore alternato per leggibilità
    vals = [mu, W, round(Ts, 6), round(Tm, 8)]
    fmts = [None, None, num_fmt_s, num_fmt_m]
    for col, (v, f) in enumerate(zip(vals, fmts), start=1):
        cell = ws.cell(row=row_idx, column=col, value=v)
        cell.alignment = center; cell.border = border
        if f: cell.number_format = f # Applica formato numerico (decimali)
        if fill: cell.fill = fill

# Blocca la prima riga (intestazione) durante lo scroll
ws.freeze_panes = 'A2'

# --- 5. FOGLIO INFORMATIVO (README) ---
ws2 = wb.create_sheet("README")
ws2['A1'] = 'DATASET MONZA - Sweep mu e W'
ws2['A1'].font = Font(bold=True, size=14)
ws2['A2'] = f'mu range: {mu_vals[0]} ÷ {mu_vals[-1]} (step 0.2)  —  {len(mu_vals)} valori'
ws2['A3'] = f'W  range: {W_vals[0]} ÷ {W_vals[-1]} (step 0.2)  —  {len(W_vals)} valori'
ws2['A4'] = f'Totale combinazioni: {len(results)}'
ws2['A6'] = 'Colonne:'
ws2['A7'] = '  mu [-]      → coefficiente di aderenza pneumatico'
ws2['A8'] = '  W [s/L]     → peso del consumo nella funzione obiettivo'
ws2['A9'] = '  Tempo [s]   → tempo giro ottimizzato (secondi)'
ws2['A10']= '  Tempo [min] → tempo giro ottimizzato (minuti)'
ws2.column_dimensions['A'].width = 55

# Salvataggio finale del file
out_path = "monza_dataset.xlsx"
wb.save(out_path)
print(f"\nDataset salvato in: {out_path}")
print(f"Righe dati: {len(results)}  (+ 1 header)")